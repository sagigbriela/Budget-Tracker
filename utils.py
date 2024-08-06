import openpyxl
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import json
import os

def read_json(file_path):
    try:
        with open(file_path, 'r') as file:
            data = json.load(file)
            return data["balance"], data["expense_list"]
    except (FileNotFoundError, json.JSONDecodeError):
        return 0, []

def save_budget_detail(file_path, balance, expense_list):
    data = {
        "balance" : balance,
        "expense_list" : expense_list
    }
    with open(file_path, 'w') as file:
        json.dump(data, file, indent=4)

def add_expense(balance, expense_list):
    expense_desc = input("Enter expense description: ")
    expense_amount = float(input("Enter expense amount: "))
    if expense_amount > balance:
        print("Error.\nExpense amount is over the balance budget")
        return 0
    
    expense_list.append({
        "expense_desc" : expense_desc,
        "expense_amount" : expense_amount
    })
    print(f"Added expense: {expense_desc}, Amount: {expense_amount}")
    return balance - expense_amount

def budget_detail(balance, expense_list):
    print(f"Total budget: {balance}")
    total_spend = 0
    for expense in expense_list:
        print(f"- {expense["expense_desc"]}: {expense["expense_amount"]}")
        total_spend += expense["expense_amount"]
    print(f"Total Spent: {total_spend}")
    print(f"Remaining budget: ", (balance - total_spend))
    return (balance - total_spend)

def style_cell(cell, font=None, fill=None, border=None):
    """
    Apply styles to a single cell.
    
    Parameters:
    cell (openpyxl.cell.cell.Cell): The cell to style.
    font (openpyxl.styles.Font): Font style to apply.
    fill (openpyxl.styles.PatternFill): Fill style to apply.
    border (openpyxl.styles.Border): Border style to apply.
    """
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if border:
        cell.border = border

def budget_sheet(expense_list, balance):
    #File address
    if not os.path.isdir("./output"):
        os.mkdir("output")
    template_path = 'template.xlsx'
    output_file = "./output/output.xlsx"
    wb = load_workbook(template_path)
    sheet_name = wb.sheetnames[0]
    ws = wb[sheet_name]
    
    #Add the info to lists so it can be converted as Dataframe later.
    descriptions = []
    amounts = []
    for expense in expense_list:
        try:
            descriptions.append(expense["expense_desc"])
            amounts.append(expense["expense_amount"])
        except KeyError as e:
            print(f"Missing key: {e} in expenses: {expense}")
    df = pd.DataFrame({"Amount": amounts}, index=descriptions)
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    border = Border(left=Side(style='thin'), 
                    right=Side(style='thin'), 
                    top=Side(style='thin'), 
                    bottom=Side(style='thin'))


    #Starting position in the template
    start_row = 6
    start_col = 1
    
    #Write Date and Buget
    ws["A4"] = f"Initial Budget:  {balance}"
    ws["A4"].font = Font(bold=True)
    
    #Write header manually
    headers = ["Description", "Amount"]
    for c_idx, header in enumerate(headers, start=start_col):
        cell = ws.cell(row=start_row, column=c_idx, value=header)
        style_cell(cell, font=header_font, fill=header_fill, border=border)
    
    #Write DataFrame to the sheet
    for r_idx, (index, row) in enumerate(df.iterrows(), start=start_row + 1):
        ws.cell(row=r_idx, column=start_col, value=index)
        ws.cell(row=r_idx, column=start_col + 1, value=row["Amount"])
        
        #Apply styles to data rows
        for c_idx in range(start_col, start_col + len(headers)):
            cell = ws.cell(row=r_idx, column=c_idx)
            style_cell(cell, border=border)
    
    len_df = len(descriptions)
    cell_total = str(6 + len_df + 1)
    ws["A"+cell_total] = "Total"
    ws["B"+cell_total] = budget_detail(balance, expense_list)

    #Save the updated workbook
    wb.save(output_file)
    print(f"DataFrame pasted into template and saved to {output_file}")